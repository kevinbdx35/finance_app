import os
import io
import csv
import json
import sys
import time
import secrets
import subprocess
import urllib.request
from datetime import datetime, date
from flask import (
    Flask, render_template, request, redirect, url_for,
    jsonify, send_file, flash, abort
)
from werkzeug.utils import secure_filename

import database as db
import pdf_generator as pdf

app = Flask(__name__)

# Clé secrète persistante générée automatiquement au premier lancement
_KEY_FILE = os.path.join(os.path.dirname(__file__), '.secret_key')
if os.path.exists(_KEY_FILE):
    with open(_KEY_FILE) as _f:
        app.secret_key = _f.read().strip()
else:
    app.secret_key = secrets.token_hex(32)
    with open(_KEY_FILE, 'w') as _f:
        _f.write(app.secret_key)

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
DOCUMENTS_FOLDER = os.path.join(os.path.dirname(__file__), 'documents')
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'webp'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DOCUMENTS_FOLDER, exist_ok=True)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


MAGIC_BYTES = {
    b'%PDF': 'pdf',
    b'\x89PNG': 'png',
    b'\xff\xd8': 'jpg',
    b'GIF8': 'gif',
    b'RIFF': 'webp',
}

def valid_file_content(file):
    header = file.read(8)
    file.seek(0)
    for magic, _ in MAGIC_BYTES.items():
        if header[:len(magic)] == magic:
            return True
    return False


def get_currency():
    settings = db.get_settings()
    return settings.get('currency', '€')


GITHUB_VERSION_URL = (
    'https://raw.githubusercontent.com/kevinbdx35/finance_app/main/VERSION'
)
_update_cache = {'checked_at': 0.0, 'remote_version': None, 'error': None}
_UPDATE_TTL = 600  # secondes entre deux vérifications (10 min)


def get_version():
    try:
        version_file = os.path.join(os.path.dirname(__file__), 'VERSION')
        with open(version_file) as f:
            return f.read().strip()
    except Exception:
        return '1.0.0'


def _parse_version(v):
    try:
        return tuple(int(x) for x in v.strip().split('.'))
    except Exception:
        return (0, 0, 0)


@app.context_processor
def inject_globals():
    settings = db.get_settings()
    balance = db.get_total_balance()
    last_update = db.get_last_update()
    return dict(
        settings=settings,
        sidebar_balance=balance,
        sidebar_last_update=last_update,
        app_version=get_version(),
    )


# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.route('/')
def dashboard():
    today = date.today()
    year, month = today.year, today.month

    # Previous month
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1

    income_cur, expense_cur = db.get_monthly_stats(year, month)
    income_prev, expense_prev = db.get_monthly_stats(prev_year, prev_month)
    balance = db.get_total_balance()

    monthly_balance = income_cur - expense_cur
    monthly_balance_prev = income_prev - expense_prev

    chart_data = db.get_monthly_chart_data(12)
    donut_data = db.get_category_donut_data(year, month)
    last_transactions = db.get_last_transactions(5)
    budget_alerts = db.get_budget_alerts(year, month)
    categories = db.get_categories()

    def delta(cur, prev):
        if prev == 0:
            return None
        return round((cur - prev) / prev * 100, 1)

    return render_template('dashboard.html',
        page='dashboard',
        balance=balance,
        income_cur=income_cur,
        expense_cur=expense_cur,
        monthly_balance=monthly_balance,
        delta_income=delta(income_cur, income_prev),
        delta_expense=delta(expense_cur, expense_prev),
        delta_balance=delta(monthly_balance, monthly_balance_prev),
        chart_data=json.dumps(chart_data),
        donut_data=json.dumps(donut_data),
        last_transactions=last_transactions,
        budget_alerts=budget_alerts,
        categories=categories,
        currency=get_currency(),
        current_month=f"{month:02d}/{year}",
    )


# ── Transactions ──────────────────────────────────────────────────────────────

@app.route('/transactions')
def transactions():
    try:
        page = max(1, int(request.args.get('page', 1)))
    except (ValueError, TypeError):
        page = 1
    filters = {
        'type': request.args.get('type', ''),
        'category_id': request.args.get('category_id', ''),
        'date_from': request.args.get('date_from', ''),
        'date_to': request.args.get('date_to', ''),
        'amount_min': request.args.get('amount_min', ''),
        'amount_max': request.args.get('amount_max', ''),
        'search': request.args.get('search', ''),
        'sort': request.args.get('sort', 'date'),
        'dir': request.args.get('dir', 'desc'),
    }
    # Clean empties
    clean_filters = {k: v for k, v in filters.items() if v}
    if 'amount_min' in clean_filters:
        try:
            clean_filters['amount_min'] = float(clean_filters['amount_min'])
        except ValueError:
            del clean_filters['amount_min']
    if 'amount_max' in clean_filters:
        try:
            clean_filters['amount_max'] = float(clean_filters['amount_max'])
        except ValueError:
            del clean_filters['amount_max']

    txs, total = db.get_transactions(clean_filters, page=page, per_page=20)
    categories = db.get_categories()
    total_pages = max(1, (total + 19) // 20)

    return render_template('transactions.html',
        page='transactions',
        transactions=txs,
        categories=categories,
        total=total,
        current_page=page,
        total_pages=total_pages,
        filters=filters,
        currency=get_currency(),
    )


@app.route('/transactions/add', methods=['POST'])
def add_transaction():
    data = request.form
    attachment_path = None
    file = request.files.get('attachment')
    if file and file.filename and allowed_file(file.filename):
        if not valid_file_content(file):
            flash('Fichier invalide ou corrompu.', 'error')
            return redirect(url_for('transactions'))
        filename = secure_filename(file.filename)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S_')
        filename = ts + filename
        file.save(os.path.join(UPLOAD_FOLDER, filename))
        attachment_path = filename

    try:
        amount = float(data['amount'])
        if amount < 0:
            flash('Le montant ne peut pas être négatif.', 'error')
            return redirect(url_for('transactions'))
        db.create_transaction(
            date=data['date'],
            label=data['label'][:200],
            amount=amount,
            typ=data['type'],
            category_id=data.get('category_id') or None,
            notes=data.get('notes', '')[:500],
            attachment_path=attachment_path,
        )
        flash('Transaction ajoutée.', 'success')
    except Exception as e:
        flash(f'Erreur : {e}', 'error')

    return redirect(url_for('transactions'))


@app.route('/transactions/<int:tx_id>/edit', methods=['POST'])
def edit_transaction(tx_id):
    tx = db.get_transaction(tx_id)
    if not tx:
        abort(404)

    data = request.form
    attachment_path = tx['attachment_path']

    file = request.files.get('attachment')
    if file and file.filename and allowed_file(file.filename):
        if not valid_file_content(file):
            flash('Fichier invalide ou corrompu.', 'error')
            return redirect(url_for('transactions'))
        if attachment_path:
            old = os.path.join(UPLOAD_FOLDER, attachment_path)
            if os.path.exists(old):
                os.remove(old)
        filename = secure_filename(file.filename)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S_')
        filename = ts + filename
        file.save(os.path.join(UPLOAD_FOLDER, filename))
        attachment_path = filename

    # Remove attachment if requested
    if data.get('remove_attachment') == '1' and attachment_path:
        old = os.path.join(UPLOAD_FOLDER, attachment_path)
        if os.path.exists(old):
            os.remove(old)
        attachment_path = None

    try:
        amount = float(data['amount'])
        if amount < 0:
            flash('Le montant ne peut pas être négatif.', 'error')
            return redirect(url_for('transactions'))
        db.update_transaction(
            tx_id=tx_id,
            date=data['date'],
            label=data['label'][:200],
            amount=amount,
            typ=data['type'],
            category_id=data.get('category_id') or None,
            notes=data.get('notes', '')[:500],
            attachment_path=attachment_path,
        )
        flash('Transaction mise à jour.', 'success')
    except Exception as e:
        flash(f'Erreur : {e}', 'error')

    return redirect(url_for('transactions'))


@app.route('/transactions/<int:tx_id>/delete', methods=['POST'])
def delete_transaction(tx_id):
    attachment = db.delete_transaction(tx_id)
    if attachment:
        path = os.path.join(UPLOAD_FOLDER, attachment)
        if os.path.exists(path):
            os.remove(path)
    flash('Transaction supprimée.', 'success')
    return redirect(url_for('transactions'))


@app.route('/transactions/<int:tx_id>/attachment')
def view_attachment(tx_id):
    tx = db.get_transaction(tx_id)
    if not tx or not tx['attachment_path']:
        abort(404)
    real_folder = os.path.realpath(UPLOAD_FOLDER)
    safe_path = os.path.realpath(os.path.join(UPLOAD_FOLDER, tx['attachment_path']))
    if not safe_path.startswith(real_folder + os.sep):
        app.logger.warning(f"Path traversal bloqué : {tx['attachment_path']}")
        abort(404)
    return send_file(safe_path)


def _extract_filters(args):
    filters = {}
    for key in ('type', 'category_id', 'date_from', 'date_to', 'search', 'sort', 'dir'):
        if args.get(key):
            filters[key] = args[key]
    for key in ('amount_min', 'amount_max'):
        try:
            v = float(args[key])
            filters[key] = v
        except (KeyError, TypeError, ValueError):
            pass
    return filters


@app.route('/transactions/export/csv')
def export_csv():
    clean_filters = _extract_filters(request.args)
    txs, _ = db.get_transactions(clean_filters, page=1, per_page=99999)
    currency = get_currency()

    output = io.StringIO()
    output.write('\ufeff')  # BOM UTF-8
    writer = csv.writer(output, delimiter=';')
    writer.writerow(['Date', 'Libellé', 'Type', 'Catégorie', 'Montant', 'Notes'])
    for tx in txs:
        writer.writerow([
            tx['date'],
            tx['label'],
            'Entrée' if tx['type'] == 'income' else 'Sortie',
            tx.get('category_name') or '',
            f"{tx['amount']:.2f}",
            tx.get('notes') or '',
        ])

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f"slamm_transactions_{date.today()}.csv",
    )


@app.route('/transactions/export/excel')
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    clean_filters = _extract_filters(request.args)
    txs, _ = db.get_transactions(clean_filters, page=1, per_page=99999)
    currency = get_currency()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    header_fill = PatternFill('solid', fgColor='C8102E')
    header_font = Font(color='FFFFFF', bold=True, name='Calibri', size=11)
    green_font = Font(color='22C55E', name='Calibri', size=10)
    red_font = Font(color='EF4444', name='Calibri', size=10)
    bold_font = Font(bold=True, name='Calibri', size=11)

    headers = ['Date', 'Libellé', 'Type', 'Catégorie', f'Montant ({currency})', 'Notes']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, tx in enumerate(txs, 2):
        ws.cell(row=row_idx, column=1, value=tx['date'])
        ws.cell(row=row_idx, column=2, value=tx['label'])
        typ_label = 'Entrée' if tx['type'] == 'income' else 'Sortie'
        cell_type = ws.cell(row=row_idx, column=3, value=typ_label)
        cell_type.font = green_font if tx['type'] == 'income' else red_font
        ws.cell(row=row_idx, column=4, value=tx.get('category_name') or '')
        amount_cell = ws.cell(row=row_idx, column=5, value=tx['amount'])
        amount_cell.font = green_font if tx['type'] == 'income' else red_font
        amount_cell.number_format = f'#,##0.00 "{currency}"'
        ws.cell(row=row_idx, column=6, value=tx.get('notes') or '')

    # Totals
    last_row = len(txs) + 2
    ws.cell(row=last_row, column=4, value='TOTAL Entrées').font = bold_font
    income_total = sum(t['amount'] for t in txs if t['type'] == 'income')
    expense_total = sum(t['amount'] for t in txs if t['type'] == 'expense')
    t_income = ws.cell(row=last_row, column=5, value=income_total)
    t_income.font = Font(bold=True, color='22C55E', name='Calibri')
    t_income.number_format = f'#,##0.00 "{currency}"'

    ws.cell(row=last_row + 1, column=4, value='TOTAL Sorties').font = bold_font
    t_expense = ws.cell(row=last_row + 1, column=5, value=expense_total)
    t_expense.font = Font(bold=True, color='EF4444', name='Calibri')
    t_expense.number_format = f'#,##0.00 "{currency}"'

    ws.cell(row=last_row + 2, column=4, value='SOLDE').font = bold_font
    balance = income_total - expense_total
    t_balance = ws.cell(row=last_row + 2, column=5, value=balance)
    t_balance.font = Font(bold=True, color='22C55E' if balance >= 0 else 'EF4444', name='Calibri')
    t_balance.number_format = f'#,##0.00 "{currency}"'

    # Auto-width
    for col in range(1, 7):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"slamm_transactions_{date.today()}.xlsx",
    )


# ── Documents ─────────────────────────────────────────────────────────────────

@app.route('/documents')
def documents():
    search = request.args.get('search', '')
    category = request.args.get('category', '')
    docs = db.get_documents(search or None, category or None)
    return render_template('documents.html',
        page='documents',
        documents=docs,
        categories=db.DOCUMENT_CATEGORIES,
        search=search,
        active_category=category,
        today=date.today().isoformat(),
    )


@app.route('/documents/add', methods=['POST'])
def add_document():
    file = request.files.get('file')
    if not file or not file.filename:
        flash('Fichier requis.', 'error')
        return redirect(url_for('documents'))

    filename = secure_filename(file.filename)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S_')
    filename = ts + filename
    file.save(os.path.join(DOCUMENTS_FOLDER, filename))

    db.create_document(
        name=request.form.get('name', filename),
        date=request.form.get('date', date.today().isoformat()),
        category=request.form.get('category', 'Autre'),
        notes=request.form.get('notes', ''),
        file_path=filename,
    )
    flash('Document ajouté.', 'success')
    return redirect(url_for('documents'))


@app.route('/documents/<int:doc_id>/download')
def download_document(doc_id):
    doc = db.get_document(doc_id)
    if not doc:
        abort(404)
    real_folder = os.path.realpath(DOCUMENTS_FOLDER)
    safe_path = os.path.realpath(os.path.join(DOCUMENTS_FOLDER, doc['file_path']))
    if not safe_path.startswith(real_folder + os.sep):
        app.logger.warning(f"Path traversal bloqué (document) : {doc['file_path']}")
        abort(404)
    if not os.path.exists(safe_path):
        abort(404)
    return send_file(safe_path, as_attachment=True, download_name=doc['name'])


@app.route('/documents/<int:doc_id>/edit', methods=['POST'])
def edit_document(doc_id):
    doc = db.get_document(doc_id)
    if not doc:
        abort(404)
    name = request.form.get('name', '').strip() or doc['name']
    date_val = request.form.get('date') or doc['date']
    category = request.form.get('category') or doc['category']
    notes = request.form.get('notes', '')
    db.update_document(doc_id, name[:200], date_val, category, notes[:500])
    flash('Document mis à jour.', 'success')
    return redirect(url_for('documents'))


@app.route('/documents/<int:doc_id>/delete', methods=['POST'])
def delete_document(doc_id):
    file_path = db.delete_document(doc_id)
    if file_path:
        path = os.path.join(DOCUMENTS_FOLDER, file_path)
        if os.path.exists(path):
            os.remove(path)
    flash('Document supprimé.', 'success')
    return redirect(url_for('documents'))


# ── Budget ────────────────────────────────────────────────────────────────────

@app.route('/budget')
def budget():
    try:
        year = int(request.args.get('year', date.today().year))
    except (ValueError, TypeError):
        year = date.today().year
    budget_data = db.get_budget_realization(year)
    categories = db.get_categories()
    years = list(range(date.today().year - 2, date.today().year + 3))

    income_rows = [r for r in budget_data if r['type'] == 'income']
    expense_rows = [r for r in budget_data if r['type'] == 'expense']

    return render_template('budget.html',
        page='budget',
        year=year,
        years=years,
        income_rows=income_rows,
        expense_rows=expense_rows,
        categories=categories,
        currency=get_currency(),
    )


@app.route('/budget/save', methods=['POST'])
def save_budget():
    year = int(request.form.get('year', date.today().year))
    for key, value in request.form.items():
        if key.startswith('budget_'):
            # Format: budget_{category_id}_{type}
            parts = key.split('_')
            if len(parts) == 3:
                try:
                    cat_id = int(parts[1])
                    typ = parts[2]
                    amount = max(0.0, float(value or 0))
                    db.save_budget(year, cat_id, typ, amount)
                except (ValueError, IndexError):
                    pass
    flash('Budget enregistré.', 'success')
    return redirect(url_for('budget', year=year))


# ── Rapports ──────────────────────────────────────────────────────────────────

@app.route('/rapports')
def rapports():
    years = list(range(date.today().year - 3, date.today().year + 1))
    return render_template('rapports.html',
        page='rapports',
        years=years,
        current_year=date.today().year,
        current_month=date.today().month,
    )


@app.route('/rapports/monthly', methods=['POST'])
def rapport_monthly():
    try:
        year = int(request.form['year'])
        month = int(request.form['month'])
        if not (1 <= month <= 12):
            raise ValueError
    except (ValueError, KeyError):
        abort(400)
    currency = get_currency()
    txs = db.get_transactions_for_report(year, month)
    buf = pdf.generate_monthly_report(txs, year, month, currency)
    MONTHS_FR = ['jan', 'fev', 'mar', 'avr', 'mai', 'jun',
                 'jul', 'aou', 'sep', 'oct', 'nov', 'dec']
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f"slamm_rapport_{MONTHS_FR[month-1]}_{year}.pdf")


@app.route('/rapports/annual', methods=['POST'])
def rapport_annual():
    try:
        year = int(request.form['year'])
    except (ValueError, KeyError):
        abort(400)
    currency = get_currency()
    txs = db.get_transactions_for_report(year)
    budget_data = db.get_budget_realization(year)
    buf = pdf.generate_annual_report(txs, budget_data, year, currency)
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f"slamm_rapport_annuel_{year}.pdf")


@app.route('/rapports/receipt/<int:tx_id>')
def rapport_receipt(tx_id):
    tx = db.get_transaction(tx_id)
    if not tx:
        abort(404)
    currency = get_currency()
    buf = pdf.generate_receipt(tx, currency)
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f"slamm_recu_{tx_id}.pdf")


# ── Catégories ────────────────────────────────────────────────────────────────

@app.route('/categories')
def categories():
    cats = db.get_categories()
    usage = db.get_categories_usage()
    return render_template('categories.html',
        page='categories',
        categories=cats,
        usage=usage,
    )


@app.route('/categories/add', methods=['POST'])
def add_category():
    name = request.form.get('name', '').strip()
    typ = request.form.get('type', 'both')
    color = request.form.get('color', '#888888')
    if not name:
        flash('Nom requis.', 'error')
        return redirect(url_for('categories'))
    try:
        db.create_category(name, typ, color)
        flash('Catégorie créée.', 'success')
    except Exception as e:
        flash(f'Erreur : {e}', 'error')
    return redirect(url_for('categories'))


@app.route('/categories/<int:cat_id>/edit', methods=['POST'])
def edit_category(cat_id):
    name = request.form.get('name', '').strip()
    typ = request.form.get('type', 'both')
    color = request.form.get('color', '#888888')
    if not name:
        flash('Nom requis.', 'error')
        return redirect(url_for('categories'))
    db.update_category(cat_id, name, typ, color)
    flash('Catégorie mise à jour.', 'success')
    return redirect(url_for('categories'))


@app.route('/categories/<int:cat_id>/delete', methods=['POST'])
def delete_category(cat_id):
    ok, msg = db.delete_category(cat_id)
    flash(msg, 'success' if ok else 'error')
    return redirect(url_for('categories'))


# ── Paramètres ────────────────────────────────────────────────────────────────

@app.route('/parametres', methods=['GET', 'POST'])
def parametres():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'save_settings':
            db.update_setting('association_name', request.form.get('association_name', 'SLAMM'))
            db.update_setting('location', request.form.get('location', ''))
            db.update_setting('currency', request.form.get('currency', '€'))
            flash('Paramètres enregistrés.', 'success')
        elif action == 'export_db':
            db_path = db.DB_PATH
            return send_file(db_path, as_attachment=True,
                             download_name='slamm_finances.db',
                             mimetype='application/octet-stream')
        elif action == 'import_db':
            file = request.files.get('db_file')
            if file and file.filename.endswith('.db'):
                header = file.read(16)
                file.seek(0)
                if not header.startswith(b'SQLite format 3'):
                    app.logger.warning("Tentative d'import d'un fichier non-SQLite rejetée.")
                    flash("Fichier invalide — ce n'est pas une base SQLite.", 'error')
                else:
                    file.save(db.DB_PATH)
                    flash("Base de données restaurée. Redémarrez l'application.", 'success')
            else:
                flash('Fichier .db requis.', 'error')
        return redirect(url_for('parametres'))

    settings = db.get_settings()
    return render_template('parametres.html',
        page='parametres',
        settings=settings,
    )


# ── API JSON ──────────────────────────────────────────────────────────────────

@app.route('/api/transaction/<int:tx_id>')
def api_transaction(tx_id):
    tx = db.get_transaction(tx_id)
    if not tx:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(tx)


@app.route('/api/document/<int:doc_id>')
def api_document(doc_id):
    doc = db.get_document(doc_id)
    if not doc:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(doc)


# ── Mise à jour automatique ───────────────────────────────────────────────────

@app.route('/api/check-update')
def api_check_update():
    """Vérifie si une nouvelle version est disponible sur GitHub (cache 10 min)."""
    now = time.time()
    if now - _update_cache['checked_at'] > _UPDATE_TTL:
        try:
            req = urllib.request.Request(
                GITHUB_VERSION_URL,
                headers={'Cache-Control': 'no-cache'},
            )
            with urllib.request.urlopen(req, timeout=5) as resp:
                _update_cache['remote_version'] = resp.read().decode().strip()
            _update_cache['error'] = None
        except Exception as e:
            _update_cache['error'] = str(e)
        _update_cache['checked_at'] = now

    local = get_version()
    remote = _update_cache['remote_version']
    update_available = bool(
        remote and _parse_version(remote) > _parse_version(local)
    )
    return jsonify({
        'local_version': local,
        'remote_version': remote,
        'update_available': update_available,
        'error': _update_cache.get('error'),
    })


@app.route('/admin/update', methods=['POST'])
def admin_update():
    """Exécute git pull puis redémarre l'application."""
    repo_dir = os.path.dirname(os.path.abspath(__file__))
    try:
        output = subprocess.check_output(
            ['git', 'pull'],
            cwd=repo_dir,
            stderr=subprocess.STDOUT,
            text=True,
            timeout=30,
        )
    except subprocess.TimeoutExpired:
        return jsonify({'success': False, 'error': 'git pull a expiré (timeout 30s).'}), 500
    except subprocess.CalledProcessError as e:
        return jsonify({'success': False, 'error': e.output}), 500

    # Invalide le cache de version pour relire le nouveau fichier
    _update_cache['checked_at'] = 0.0

    # Redémarre le process Python dans 1 seconde (laisse le temps de répondre)
    import threading
    def _restart():
        time.sleep(1)
        os.execv(sys.executable, [sys.executable] + sys.argv)
    threading.Thread(target=_restart, daemon=True).start()

    return jsonify({'success': True, 'output': output})


# ── Erreurs personnalisées ────────────────────────────────────────────────────

@app.errorhandler(404)
def page_not_found(e):
    return render_template('error.html', code=404,
                           message="Page introuvable."), 404


@app.errorhandler(500)
def server_error(e):
    app.logger.error(f"Erreur serveur : {e}")
    return render_template('error.html', code=500,
                           message="Erreur interne du serveur."), 500


if __name__ == '__main__':
    import logging
    logging.basicConfig(
        filename=os.path.join(os.path.dirname(__file__), 'slamm.log'),
        level=logging.WARNING,
        format='%(asctime)s %(levelname)s %(message)s',
    )
    try:
        db.init_db()
    except Exception as e:
        print(f"ERREUR : impossible d'initialiser la base de données : {e}")
        raise SystemExit(1)
    port = int(os.environ.get('FLASK_PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
